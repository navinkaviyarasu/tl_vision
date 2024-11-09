#!/usr/bin/env python3


import rospy
import openpyxl

from datetime import datetime
from nav_msgs.msg import Odometry
from vision.msg import VioState
from sensor_msgs.msg import Imu
from geometry_msgs.msg import PoseStamped


def generate_timestamped_filename():
    timestamp = datetime.now().strftime("%Y%m%d%H%M")
    return f"logger{timestamp}.xlsx"

def create_excel_file(filename):
    
    wb = openpyxl.Workbook()

    ws_vicon = wb.active
    ws_vicon.title = "vicon"
    ws_vicon.append(["Time", "Sequence", "Stamp", "Frame_ID", "Position_x", "Position_y",
               "Position_Z", "Orientation_x", "Orientation_y", "Orientation_z",
               "Orientation_w"])
    
    ws_vioodom = wb.create_sheet("vio")
    ws_vioodom.append(["Time", "Sequence", "Stamp", "Frame_ID", "Child_Frame_ID"
                       "Position_x", "Position_y", "Position_z", "Orientation_x",
                       "Orientation_y", "Orientation_z", "Orientation_w", "L.Velocity_x",
                       "L.Velocity_y", "L.Velocity_z", "A.Velocity_x", "A.Velocity_y", 
                       "A.Velocity_z"])
    
    ws_ekf = wb.create_sheet("ekf")
    ws_ekf.append(["Time", "Sequence", "Stamp", "Frame_ID", "Position_x", "Position_y",
                   "Position_z", "Orientation_x", "Orientation_y", "Orientation_z"
                   "Orientation_w"])


    ws_vio_state = wb.create_sheet("vio_state")
    ws_vio_state.append(["Time", "Sequence", "Timestamp", "Frame ID", 
                   "Vision Failure", "Inertial Failure", "Failure Drift", 
                   "VIO Failure", "Reset Counter"])

    # ws_imu = wb.create_sheet("IMU Data")
    # ws_imu.append(["Time", "Sequence", "Timestamp", "Orientation X", 
    #                "Orientation Y", "Orientation Z", "Orientation W"])


    wb.save(filename)
    rospy.loginfo(f"Created new Excel file with sheets: {filename}")

def vicon_callback(msg, wb):
    ws = wb["vicon"]
    ws.append([rospy.get_time(), msg.header.seq, msg.header.stamp, msg.header.frame_id, msg.pose.position.x,
               msg.pose.position.y, msg.pose.position.z, msg.pose.orientation.x, msg.pose.orientation.y,
               msg.pose.orientation.z, msg.orientation.w])
    wb.save(filename)

def odometry_callback(msg, wb):
    ws = wb["vio"]
    ws.append([rospy.get_time(), msg.header.seq, msg.header.stamp, msg.header.frame_id, msg.child_frame_id,
               msg.pose.pose.position.x, msg.pose.pose.position.y, msg.pose.pose.position.z,
               msg.pose.pose.orientation.x, msg.pose.pose.orientation.y, msg.pose.pose.orientation.z,
               msg.pose.pose.orientation.w, msg.twist.twist.linear.x, msg.twist.twist.linear.y,
               msg.twist.twist.linear.z, msg.twist.twist.angular.x, msg.twist.twist.angular.y,
               msg.twist.twist.angular.z])
    wb.save(filename)

def ekf_callback(msg, wb):
    ws = wb["ekf"]
    ws.append([rospy.get_time(), msg.header.seq, msg.header.stamp, msg.header.frame_id, msg.pose.position.x,
               msg.pose.position.y, msg.pose.position.z, msg.pose.orientation.x, msg.pose.orientation.y,
               msg.pose.orientation.z, msg.orientation.w])
    wb.save(filename)

def viostate_callback(msg, wb):
    ws = wb["vio_state"]
    ws.append([rospy.get_time(), msg.header.seq, msg.header.stamp, 
               msg.header.frame_id, msg.vision_failure, msg.inertial_failure, 
               msg.failure_drift, msg.vio_failure, msg.reset_counter])
    wb.save(filename)

# def imu_callback(msg, wb):
#     ws = wb["IMU Data"]
#     ws.append([rospy.get_time(), msg.header.seq, msg.header.stamp, 
#                msg.orientation.x, msg.orientation.y, msg.orientation.z, msg.orientation.w])
#     wb.save(filename)


def listener():
    rospy.init_node('logger', anonymous=False)


    log_filename = generate_timestamped_filename()

    wb = openpyxl.Workbook()
    create_excel_file(log_filename)

    # rospy.Subscriber('/vrpn_client_node/Akira/pose', PoseStamped, lambda msg: vicon_callback(msg, wb))
    # rospy.Subscriber('S0/basalt/odom', Odometry, lambda msg: odometry_callback(msg, wb))
    # rospy.Subscriber('/mavros/local_position/pose', PoseStamped, lambda msg: ekf_callback(msg, wb))
    rospy.Subscriber('/vision/vio_state', VioState, lambda msg: viostate_callback(msg, wb))  
    # rospy.Subscriber('/sensor/imu_data', Imu, lambda msg: imu_callback(msg, wb))

    rospy.spin()

if __name__ == '__main__':
    rospy.loginfo("Starting the logger node...")
    listener()
    rospy.loginfo("Logging in process...")